"""
generate.py — Museum Tuanku Imam Bonjol QR & Data Generator
Membaca Excel master (multi-sheet), menghasilkan:
  - data/koleksi.json     : detail tiap koleksi, key = Kode TIB
  - data/qr_mapping.json  : daftar Kode TIB per Kode QR (per box)
  - qr/<KodeQR>.png       : gambar QR code, berisi link ke website
  - laporan_validasi.txt  : daftar baris bermasalah (data kosong/duplikat)

Cara pakai:
    pip install openpyxl qrcode[pil]
    python generate.py

Konfigurasi ada di bagian CONFIG di bawah ini.
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook

# ============ CONFIG — SESUAIKAN INI ============
EXCEL_PATH = "Data_Koleksi_Museum_TIB.xlsx"
# Sheet yang berisi data koleksi (tambahkan nama sheet baru di sini kalau ada ruangan baru)
DATA_SHEETS = ["Data Koleksi Senjata", "Data Koleksi Seni Rupa"]
# Ganti dengan URL GitHub Pages asli kalian, contoh:
#   "https://namamuseum.github.io/koleksi-tib"
BASE_URL = "https://USERNAME.github.io/REPO-NAME"
OUTPUT_DIR = Path(".")
# ==================================================

REQUIRED_FIELDS = ["Kode TIB", "Kode QR", "Nama Koleksi (Umum)"]


def read_sheet(ws):
    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in r):
            continue
        row = {h: (r[i] if i < len(r) else None) for h, i in col.items()}
        rows.append(row)
    return rows


def load_all_rows(excel_path, sheet_names):
    wb = load_workbook(excel_path, data_only=True)
    all_rows = []
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"  [peringatan] sheet '{sheet_name}' tidak ditemukan, dilewati")
            continue
        rows = read_sheet(wb[sheet_name])
        for row in rows:
            row["_sheet"] = sheet_name
        all_rows.extend(rows)
    return all_rows


def build_koleksi_record(row):
    def g(key):
        v = row.get(key)
        return "" if v is None else str(v).strip()

    return {
        "kode_tib": g("Kode TIB"),
        "kode_qr": g("Kode QR"),
        "nomor_box": g("Nomor Box"),
        "nama_umum": g("Nama Koleksi (Umum)"),
        "nama_daerah": g("Nama Koleksi (Daerah)"),
        "no_registrasi": g("No. Registrasi"),
        "no_inventaris_b": g("No. Inventaris B"),
        "no_inventaris_l": g("No. Inventaris L"),
        "jenis_koleksi": g("Jenis Koleksi"),
        "sub_jenis_koleksi": g("Sub Jenis Koleksi"),
        "bahan": g("Bahan"),
        "cara_didapat": g("Cara Didapat"),
        "tanggal_didapat": g("Tanggal Didapat"),
        "kondisi_benda": g("Kondisi Benda"),
        "tempat_asal_didapat": g("Tempat Asal Didapat"),
        "tempat_asal_dibuat": g("Tempat Asal Dibuat"),
        "ukuran": {
            "panjang": g("Panjang"), "lebar": g("Lebar"),
            "tinggi": g("Tinggi"), "tebal": g("Tebal"),
            "diameter": g("Diameter"),
        },
        "deskripsi": g("Deskripsi/Uraian"),
        "pencatat": g("Pencatat/Pengolah Data"),
        # konvensi nama file foto: images/<KodeTIB>.jpg — tinggal ditaruh filenya
        "foto": f"images/{g('Kode TIB')}.jpg" if g("Kode TIB") else "",
    }


def main():
    print(f"Membaca {EXCEL_PATH} ...")
    rows = load_all_rows(EXCEL_PATH, DATA_SHEETS)
    print(f"  {len(rows)} baris ditemukan dari sheet: {', '.join(DATA_SHEETS)}")

    koleksi = {}
    qr_mapping = {}
    problems = []
    seen_tib = set()

    for row in rows:
        kode_tib = (row.get("Kode TIB") or "").strip()
        kode_qr = (row.get("Kode QR") or "").strip()
        nama = (row.get("Nama Koleksi (Umum)") or "").strip()
        sheet = row.get("_sheet")

        missing = [f for f in REQUIRED_FIELDS if not (row.get(f) or "").strip()]
        if missing:
            problems.append(f"[{sheet}] '{nama or '(tanpa nama)'}' — field kosong: {', '.join(missing)} — DILEWATI, tidak masuk json/QR")
            continue

        if kode_tib in seen_tib:
            problems.append(f"[{sheet}] Kode TIB duplikat: {kode_tib} ({nama}) — cek data")
        seen_tib.add(kode_tib)

        koleksi[kode_tib] = build_koleksi_record(row)
        qr_mapping.setdefault(kode_qr, []).append(kode_tib)

    # --- simpan JSON ---
    data_dir = OUTPUT_DIR / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    with open(data_dir / "koleksi.json", "w", encoding="utf-8") as f:
        json.dump(koleksi, f, ensure_ascii=False, indent=2)
    with open(data_dir / "qr_mapping.json", "w", encoding="utf-8") as f:
        json.dump(qr_mapping, f, ensure_ascii=False, indent=2)
    print(f"  -> data/koleksi.json ({len(koleksi)} koleksi)")
    print(f"  -> data/qr_mapping.json ({len(qr_mapping)} QR)")

    # --- generate QR code images ---
    try:
        import qrcode
        qr_dir = OUTPUT_DIR / "qr"
        qr_dir.mkdir(parents=True, exist_ok=True)
        for kode_qr in qr_mapping:
            url = f"{BASE_URL}/qr/{kode_qr}"
            img = qrcode.make(url, box_size=10, border=3)
            img.save(qr_dir / f"{kode_qr}.png")
        print(f"  -> {len(qr_mapping)} file QR disimpan di qr/")
    except ImportError:
        print("  [peringatan] library 'qrcode' belum terpasang.")
        print("  Jalankan: pip install qrcode[pil]   lalu ulangi script ini untuk membuat gambar QR-nya.")

    # --- laporan validasi ---
    report_path = OUTPUT_DIR / "laporan_validasi.txt"
    with open(report_path, "w", encoding="utf-8") as f:
        if problems:
            f.write(f"Ditemukan {len(problems)} isu:\n\n")
            for p in problems:
                f.write(f"- {p}\n")
        else:
            f.write("Tidak ada isu ditemukan. Semua data valid.\n")
    print(f"  -> laporan_validasi.txt ({len(problems)} isu)")

    print("\nSelesai.")


if __name__ == "__main__":
    main()
